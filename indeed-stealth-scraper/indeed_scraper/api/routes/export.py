"""
routes/export.py — GET /api/export/excel, /api/export/pdf, /api/export/pdf/{job_id}
"""

import io
from typing import Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from indeed_scraper.api.scrape_runner import get_jobs
from indeed_scraper.job_parser import JOB_FIELDS

router = APIRouter()

# Human-readable column headers matching JOB_FIELDS order
_HEADERS = [
    "Job ID", "Title", "Company", "Location", "Salary",
    "Description", "Posted Date", "URL", "Employment Type", "Remote", "Rating",
]


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


@router.get("/export/excel")
async def export_excel() -> StreamingResponse:
    """Stream all scraped jobs as an .xlsx file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    jobs = get_jobs()
    if not jobs:
        raise HTTPException(status_code=404, detail="No jobs to export.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indeed Jobs"

    # Header row styling
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, job in enumerate(jobs, start=2):
        for col_idx, field in enumerate(JOB_FIELDS, start=1):
            val = job.get(field)
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-size columns (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=indeed_jobs.xlsx"},
    )


# ---------------------------------------------------------------------------
# PDF export — full table
# ---------------------------------------------------------------------------


@router.get("/export/pdf")
async def export_pdf() -> StreamingResponse:
    """Stream all scraped jobs as a PDF table."""
    jobs = get_jobs()
    if not jobs:
        raise HTTPException(status_code=404, detail="No jobs to export.")
    return _build_pdf_response(jobs, filename="indeed_jobs.pdf")


# ---------------------------------------------------------------------------
# PDF export — single job
# ---------------------------------------------------------------------------


@router.get("/export/pdf/{job_id}")
async def export_single_pdf(job_id: str) -> StreamingResponse:
    """Stream a single job's details as a PDF."""
    jobs = get_jobs()
    job: Optional[dict] = next((j for j in jobs if j.get("job_id") == job_id), None)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job '{job_id}' not found.")
    return _build_single_job_pdf(job)


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------


def _build_pdf_response(jobs: list[dict], filename: str) -> StreamingResponse:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed.")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()

    # Use a subset of columns for the table (description is too long)
    table_fields = ["title", "company", "location", "salary", "employment_type", "remote", "posted_date"]
    table_headers = ["Title", "Company", "Location", "Salary", "Type", "Remote", "Posted"]

    data = [table_headers]
    for job in jobs:
        row = []
        for f in table_fields:
            val = job.get(f)
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            row.append(str(val or ""))
        data.append(row)

    col_widths = [5*cm, 4*cm, 4*cm, 3.5*cm, 3*cm, 2*cm, 2.5*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#1a1a1a"), colors.HexColor("#0f0f0f")]),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e0e0e0")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#333333")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("WORDWRAP", (0, 0), (-1, -1), True),
    ]))

    title = Paragraph(f"<b>Indeed Canada — Job Results ({len(jobs)} jobs)</b>", styles["Title"])
    elements = [title, Spacer(1, 0.5*cm), table]
    doc.build(elements)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _build_single_job_pdf(job: dict) -> StreamingResponse:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import cm
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed.")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, leading=14,
                                 textColor=colors.HexColor("#e0e0e0"))
    heading_style = ParagraphStyle("heading", parent=styles["Heading1"],
                                    textColor=colors.HexColor("#3b82f6"))

    elements = []
    elements.append(Paragraph(job.get("title") or "Job Details", heading_style))
    elements.append(Spacer(1, 0.3*cm))

    meta_rows = [
        ("Company", job.get("company") or "N/A"),
        ("Location", job.get("location") or "N/A"),
        ("Salary", job.get("salary") or "N/A"),
        ("Posted", job.get("posted_date") or "N/A"),
        ("Type", job.get("employment_type") or "N/A"),
        ("Remote", "Yes" if job.get("remote") else "No"),
        ("Rating", str(job.get("company_rating") or "N/A")),
        ("URL", job.get("url") or "N/A"),
    ]
    meta_table = Table(
        [[k, v] for k, v in meta_rows],
        colWidths=[4*cm, 13*cm],
    )
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#e0e0e0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#333333")),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 0.5*cm))

    desc = job.get("description") or "No description available."
    elements.append(Paragraph("<b>Description</b>", ParagraphStyle(
        "subhead", parent=styles["Heading2"], textColor=colors.HexColor("#3b82f6"), fontSize=11)))
    elements.append(Spacer(1, 0.2*cm))
    # Escape for ReportLab XML
    desc_safe = desc.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    elements.append(Paragraph(desc_safe, body_style))

    doc.build(elements)
    buf.seek(0)

    safe_id = job.get("job_id", "job")
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=job_{safe_id}.pdf"},
    )
